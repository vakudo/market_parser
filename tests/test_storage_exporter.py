from datetime import UTC, datetime

from openpyxl import load_workbook

from market_parser.exporter import build_price_matrix, build_sheet_values, export_monthly_xlsx
from market_parser.models import ProductPrice
from market_parser.storage import Storage


def test_storage_upserts_observation_for_same_day(tmp_path) -> None:
    storage = Storage(tmp_path / "prices.sqlite")
    storage.init_db()
    run_id = storage.start_run(["detmir"])
    first = ProductPrice(
        store_slug="detmir",
        store_name="Детский мир",
        category="Детское питание",
        brand="NAN",
        product_name="Смесь NAN 800г",
        product_url="https://example.test/1",
        product_id="1",
        regular_price_kopecks=200000,
        collected_at=datetime(2026, 5, 19, 6, 0, tzinfo=UTC),
    )
    second = ProductPrice(
        store_slug="detmir",
        store_name="Детский мир",
        category="Детское питание",
        brand="NAN",
        product_name="Смесь NAN 800г",
        product_url="https://example.test/1",
        product_id="1",
        regular_price_kopecks=210000,
        collected_at=datetime(2026, 5, 19, 7, 0, tzinfo=UTC),
    )

    storage.save_prices([first], run_id=run_id)
    storage.save_prices([second], run_id=run_id)
    matrix = build_price_matrix(
        storage,
        datetime(2026, 5, 1, tzinfo=UTC).date(),
        datetime(2026, 5, 31, tzinfo=UTC).date(),
    )

    assert len(matrix.rows) == 1
    assert matrix.rows[0][5] == 2100.0


def test_export_monthly_xlsx_creates_date_blocks(tmp_path) -> None:
    storage = Storage(tmp_path / "prices.sqlite")
    storage.init_db()
    run_id = storage.start_run(["wildberries"])
    storage.save_prices(
        [
            ProductPrice(
                store_slug="wildberries",
                store_name="Вайлдберриз",
                category="Детское питание",
                brand="ФрутоНяня",
                product_name="Каша кукурузная безмолочная, 180г",
                product_url="https://example.test/wb",
                product_id="327468652",
                regular_price_kopecks=15900,
                promo_price_kopecks=10100,
                collected_at=datetime(2026, 5, 19, 6, 0, tzinfo=UTC),
            )
        ],
        run_id=run_id,
    )

    output = export_monthly_xlsx(storage, tmp_path / "exports", "2026-05")
    workbook = load_workbook(output)
    sheet = workbook["Выгрузка"]

    assert sheet["A3"].value == "Сеть"
    assert sheet["F2"].value == "19.05.2026"
    assert sheet["F3"].value == "Цена без скидки (регулярная)"
    assert sheet["G3"].value == "Цена со скидкой"
    assert sheet["A4"].value == "Вайлдберриз"
    assert sheet["F4"].value == 159.0
    assert sheet["G4"].value == 101.0
    assert "F2:H2" in [str(rng) for rng in sheet.merged_cells.ranges]


def test_export_has_rating_column_and_single_sheet(tmp_path) -> None:
    storage = Storage(tmp_path / "prices.sqlite")
    storage.init_db()
    run_id = storage.start_run(["ozon"])
    storage.save_prices(
        [
            ProductPrice(
                store_slug="ozon",
                store_name="Озон",
                category="Детское питание",
                brand="NAN",
                product_name="Детская смесь NAN, 800г",
                product_url="https://www.ozon.ru/product/1",
                product_id="1",
                regular_price_kopecks=207900,
                loyalty_price_kopecks=188000,
                rating=4.9,
                collected_at=datetime(2026, 5, 19, 6, 0, tzinfo=UTC),
            )
        ],
        run_id=run_id,
    )

    matrix = build_price_matrix(
        storage,
        datetime(2026, 5, 1, tzinfo=UTC).date(),
        datetime(2026, 5, 31, tzinfo=UTC).date(),
    )
    # base columns are now [Сеть, Бренд, Наименование, Ссылка, Рейтинг]
    assert matrix.rows[0][1] == "NAN"
    assert matrix.rows[0][3] == "https://www.ozon.ru/product/1"
    assert matrix.rows[0][4] == 4.9

    output = export_monthly_xlsx(storage, tmp_path / "exports", "2026-05")
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Выгрузка"]
    sheet = workbook["Выгрузка"]
    assert sheet["B3"].value == "Бренд"
    assert sheet["C3"].value == "Наименование товара"
    assert sheet["D3"].value == "Ссылка"
    assert sheet["E3"].value == "Рейтинг"
    assert sheet["E4"].value == 4.9
    # the link cell must be a clickable hyperlink to the product
    assert sheet["D4"].hyperlink.target == "https://www.ozon.ru/product/1"
    # the "Категория" column must be gone
    header_row = [cell.value for cell in sheet[3]]
    assert "Категория" not in header_row


def test_build_sheet_values_includes_date_row(tmp_path) -> None:
    storage = Storage(tmp_path / "prices.sqlite")
    storage.init_db()
    run_id = storage.start_run(["wildberries"])
    storage.save_prices(
        [
            ProductPrice(
                store_slug="wildberries",
                store_name="Вайлдберриз",
                category="Детское питание",
                brand="ФрутоНяня",
                product_name="Каша",
                product_url="https://example.test/wb",
                product_id="1",
                regular_price_kopecks=15900,
                collected_at=datetime(2026, 5, 19, 6, 0, tzinfo=UTC),
            )
        ],
        run_id=run_id,
    )
    matrix = build_price_matrix(
        storage,
        datetime(2026, 5, 1, tzinfo=UTC).date(),
        datetime(2026, 5, 31, tzinfo=UTC).date(),
    )

    values = build_sheet_values(matrix)

    assert values[0][5] == "19.05.2026"
    assert values[0][6] is None
    assert values[1][5] == "Цена без скидки (регулярная)"
