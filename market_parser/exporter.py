from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import PriceMatrix
from .normalize import kopecks_to_rubles
from .storage import Storage, group_observations

BASE_HEADERS = ["Сеть", "Бренд", "Наименование товара", "Рейтинг"]
PRICE_HEADERS = [
    "Цена без скидки (регулярная)",
    "Цена со скидкой",
    "Цена по карте лояльности",
]


def month_range(month: str) -> tuple[date, date]:
    start = date.fromisoformat(f"{month}-01")
    if start.month == 12:
        next_month = date(start.year + 1, 1, 1)
    else:
        next_month = date(start.year, start.month + 1, 1)
    return start, next_month - timedelta(days=1)


def last_n_days_range(days: int, today: date | None = None) -> tuple[date, date]:
    end = today or date.today()
    return end - timedelta(days=days - 1), end


def build_price_matrix(storage: Storage, start_date: date, end_date: date) -> PriceMatrix:
    dates = storage.fetch_observed_dates(start_date, end_date)
    rows = storage.fetch_price_rows(start_date, end_date)
    grouped = group_observations(rows)

    matrix_rows: list[list[Any]] = []
    for item in grouped.values():
        observations = item["observations"]
        row: list[Any] = [
            item["store_name"],
            item["brand"],
            item["name"],
            _latest_rating(observations),
        ]
        for observed_date in dates:
            obs = observations.get(observed_date.isoformat())
            if not obs:
                row.extend([None, None, None])
                continue
            row.extend(
                [
                    kopecks_to_rubles(obs["regular_price_kopecks"]),
                    kopecks_to_rubles(obs["promo_price_kopecks"]),
                    kopecks_to_rubles(obs["loyalty_price_kopecks"]),
                ]
            )
        matrix_rows.append(row)

    return PriceMatrix(dates=dates, rows=matrix_rows)


def _latest_rating(observations: dict[str, Any]) -> float | None:
    for observed_date in sorted(observations, reverse=True):
        rating = observations[observed_date].get("rating")
        if rating is not None:
            return rating
    return None


def build_sheet_values(matrix: PriceMatrix) -> list[list[Any]]:
    width = len(BASE_HEADERS) + len(matrix.dates) * len(PRICE_HEADERS)
    row_1 = [None] * len(BASE_HEADERS)
    row_2 = BASE_HEADERS.copy()
    for _ in matrix.dates:
        row_2.extend(PRICE_HEADERS)
    for observed_date in matrix.dates:
        row_1.extend([observed_date.strftime("%d.%m.%Y"), None, None])
    if len(row_1) < width:
        row_1.extend([None] * (width - len(row_1)))
    return [row_1, row_2, *matrix.rows]


def export_monthly_xlsx(storage: Storage, export_dir: Path | str, month: str) -> Path:
    start_date, end_date = month_range(month)
    matrix = build_price_matrix(storage, start_date, end_date)
    output_dir = Path(export_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{month}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Выгрузка"
    _write_matrix_sheet(sheet, matrix)
    workbook.save(output_path)
    return output_path


def _write_matrix_sheet(sheet, matrix: PriceMatrix) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(color="FFFFFF", bold=True)
    subheader_font = Font(bold=True)

    for idx, header in enumerate(BASE_HEADERS, start=1):
        cell = sheet.cell(row=3, column=idx, value=header)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    start_col = len(BASE_HEADERS) + 1
    for date_index, observed_date in enumerate(matrix.dates):
        col = start_col + date_index * len(PRICE_HEADERS)
        start_letter = get_column_letter(col)
        end_letter = get_column_letter(col + len(PRICE_HEADERS) - 1)
        sheet.merge_cells(f"{start_letter}2:{end_letter}2")
        date_cell = sheet.cell(row=2, column=col, value=observed_date.strftime("%d.%m.%Y"))
        date_cell.fill = header_fill
        date_cell.font = header_font
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        for offset, header in enumerate(PRICE_HEADERS):
            cell = sheet.cell(row=3, column=col + offset, value=header)
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rating_col = len(BASE_HEADERS)
    for row_idx, values in enumerate(matrix.rows, start=4):
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if col_idx > len(BASE_HEADERS):
                cell.number_format = '#,##0.00'
            elif col_idx == rating_col:
                cell.number_format = '0.0'
                cell.alignment = Alignment(horizontal="center")

    max_col = len(BASE_HEADERS) + len(matrix.dates) * len(PRICE_HEADERS)
    max_row = max(4, len(matrix.rows) + 3)
    sheet.freeze_panes = "E4"
    if max_col >= 1:
        sheet.auto_filter.ref = f"A3:{get_column_letter(max_col)}{max_row}"

    widths = {
        1: 22,
        2: 18,
        3: 52,
        4: 9,
    }
    for col_idx in range(1, max_col + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 16)
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 42
